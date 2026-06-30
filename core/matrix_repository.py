from __future__ import annotations

import csv
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


MATRIX_COLUMNS = [
    "m00", "m01", "m02",
    "m10", "m11", "m12",
    "m20", "m21", "m22",
]

TYPE_ALIASES = {
    "protan": "protanomaly",
    "protanopia": "protanomaly",
    "protanomalia": "protanomaly",
    "protanomaly": "protanomaly",
    "deutan": "deuteranomaly",
    "deut": "deuteranomaly",
    "deuteranopia": "deuteranomaly",
    "deuteranomalia": "deuteranomaly",
    "deuteranomaly": "deuteranomaly",
    "tritan": "tritanomaly",
    "tritanopia": "tritanomaly",
    "tritanomalia": "tritanomaly",
    "tritanomaly": "tritanomaly",
}


class MatrixRepository:
    """Carga matrices desde CSV o desde un libro Excel con una hoja por tipo."""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self._matrices: dict[str, dict[float, np.ndarray]] = {}
        self.reload()

    def reload(self) -> None:
        if not self.file_path.exists():
            raise FileNotFoundError(f"No se encontró el archivo de matrices: {self.file_path}")

        extension = self.file_path.suffix.lower()

        if extension == ".csv":
            matrices = self._load_csv()
        elif extension in (".xlsx", ".xlsm"):
            matrices = self._load_excel()
        else:
            raise ValueError("El archivo de matrices debe ser CSV o XLSX.")

        self._validate(matrices)
        self._matrices = matrices

    def get_matrix(self, tipo: str, severidad: float | int | str) -> np.ndarray:
        normalized_type = self.normalize_type(tipo)
        normalized_severity = self.normalize_severity(severidad)
        values = self._matrices[normalized_type]

        if normalized_severity in values:
            return values[normalized_severity].copy()

        # Permite valores intermedios, por ejemplo 7.3 -> 0.73.
        levels = sorted(values)
        lower = max((level for level in levels if level < normalized_severity), default=None)
        upper = min((level for level in levels if level > normalized_severity), default=None)

        if lower is None or upper is None:
            raise ValueError("La severidad solicitada está fuera del rango disponible.")

        weight = (normalized_severity - lower) / (upper - lower)
        matrix = values[lower] + ((values[upper] - values[lower]) * weight)
        return matrix.astype(np.float32)

    def get_combined_matrix(self, tipos: list[str], severidad: float | int | str) -> np.ndarray:
        if not tipos:
            raise ValueError("Debe indicar al menos un tipo de daltonismo.")

        combined = np.eye(3, dtype=np.float32)

        for tipo in tipos:
            combined = self.get_matrix(tipo, severidad) @ combined

        return combined

    def available(self) -> dict[str, list[float]]:
        return {
            tipo: sorted(levels.keys())
            for tipo, levels in self._matrices.items()
        }

    @staticmethod
    def normalize_type(tipo: str) -> str:
        if not isinstance(tipo, str):
            raise ValueError("El tipo de daltonismo debe ser texto.")

        key = tipo.strip().lower()
        key = (
            key.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
        )

        if key not in TYPE_ALIASES:
            raise ValueError(f"Tipo de daltonismo no válido: {tipo}")

        return TYPE_ALIASES[key]

    @staticmethod
    def normalize_severity(value: float | int | str) -> float:
        try:
            number = float(str(value).strip().replace(",", "."))
        except ValueError as ex:
            raise ValueError("La severidad debe ser numérica.") from ex

        if 1 < number <= 10:
            number /= 10.0

        if not 0 <= number <= 1:
            raise ValueError("La severidad debe estar entre 0 y 10, o entre 0.0 y 1.0.")

        return round(number, 4)

    def _load_csv(self) -> dict[str, dict[float, np.ndarray]]:
        matrices: dict[str, dict[float, np.ndarray]] = {}

        with self.file_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            expected = {"type", "severity", *MATRIX_COLUMNS}

            if not reader.fieldnames or not expected.issubset(reader.fieldnames):
                raise ValueError("El CSV no tiene las columnas esperadas.")

            for row in reader:
                tipo = self.normalize_type(row["type"])
                severity = self.normalize_severity(row["severity"])
                matrix = self._row_to_matrix(row)
                matrices.setdefault(tipo, {})[severity] = matrix

        return matrices

    def _load_excel(self) -> dict[str, dict[float, np.ndarray]]:
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        matrices: dict[str, dict[float, np.ndarray]] = {}

        try:
            for sheet_name in workbook.sheetnames:
                tipo = self.normalize_type(sheet_name)
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                headers = [str(value).strip() if value is not None else "" for value in next(rows)]
                positions = {name: index for index, name in enumerate(headers)}

                expected = {"severity", *MATRIX_COLUMNS}
                if not expected.issubset(positions):
                    raise ValueError(f"La hoja {sheet_name} no tiene las columnas esperadas.")

                for values in rows:
                    if not values or values[positions["severity"]] is None:
                        continue

                    severity = self.normalize_severity(values[positions["severity"]])
                    row = {column: values[positions[column]] for column in MATRIX_COLUMNS}
                    matrices.setdefault(tipo, {})[severity] = self._row_to_matrix(row)
        finally:
            workbook.close()

        return matrices

    @staticmethod
    def _row_to_matrix(row: dict) -> np.ndarray:
        values = [float(row[column]) for column in MATRIX_COLUMNS]
        return np.array(values, dtype=np.float32).reshape(3, 3)

    @staticmethod
    def _validate(matrices: dict[str, dict[float, np.ndarray]]) -> None:
        required_types = {"protanomaly", "deuteranomaly", "tritanomaly"}
        missing = required_types.difference(matrices)

        if missing:
            raise ValueError(f"Faltan matrices para: {', '.join(sorted(missing))}")

        for tipo, levels in matrices.items():
            if not levels:
                raise ValueError(f"No hay matrices cargadas para {tipo}.")

            for severity, matrix in levels.items():
                if matrix.shape != (3, 3):
                    raise ValueError(f"La matriz {tipo} {severity} no es de 3x3.")
